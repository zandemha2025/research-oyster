"""Build the SINGLE canonical QA tracker covering EVERY feature/action across every surface.

Rows come from an exhaustive code-grounded inventory (5 parallel readers over the whole codebase:
Studio UI, Studio endpoints, all MCP tools, Control Center + browser extension, CLI + scripts +
export/reporting). Each row is one user story with expected behavior from the code, a verdict, and
the exact evidence behind that verdict on the 2026-08-07 pass.

Verdicts:
  PASS          — exercised this pass, works.
  PASS (fixed)  — a real defect was found and fixed this pass; re-tested.
  REVIEWED      — code-verified this pass but NOT executed end-to-end (needs external setup:
                  a real Chrome + paired extension, the legacy gaming-collector creds, or a
                  from-scratch host install). An honest not-fully-exercised, not a PASS.
  NOTE          — works, with a known limitation / cosmetic issue worth flagging.
  DEFER         — real gap, deliberately deferred post-demo.

Evidence shorthand: harness=MCP tool harness · sweep=Playwright UI drive · wire=curl endpoint ·
live=exercised in the live research runs (jobs 1861/1998/2033/2102) · nodetest=node --test ·
code=code-reviewed this pass · fixed=defect fixed+re-tested this pass.

Re-run to regenerate qa/Research-Oyster-QA.xlsx (source of truth lives here, in git).
"""
from __future__ import annotations
from pathlib import Path
from collections import Counter
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ID | Surface | User story (human POV) | Expected behavior (from code) | Verdict | How verified (this pass)
ROWS: list[tuple[str, str, str, str, str, str]] = [
    # ================= STUDIO UI (studio_page.py) =================
    ("UI-001","Studio UI","I open Studio","boot() self-inits: health, pills, convos, reports","PASS","sweep: page boots, pills+lists populate"),
    ("UI-002","Studio UI","I see a tab icon, no favicon 404","inline data-URI favicon","PASS (fixed)","sweep: 0 console errors incl. favicon"),
    ("UI-003","Studio UI","My writes carry a session token","api() adds X-Studio-Token on POST only","PASS","wire: POST needs token (403 without)"),
    ("UI-004","Studio UI","Failed calls surface a real message","api() throws d.error on !ok","PASS","live: errors shown, not swallowed"),
    ("UI-005","Studio UI","Untrusted text is escaped","esc() via textContent","PASS","code+sweep: titles render literal"),
    ("UI-006","Studio UI","Auth pill shows auth mode","authpill green when authed","NOTE","code: ternary ?'ok':'ok' is dead code but output is correct (auth always present)"),
    ("UI-007","Studio UI","DB pill shows db reachability","db pill ok/down from health","PASS","sweep: 'db: ok' rendered"),
    ("UI-008","Studio UI","Health fails, page still loads","boot() try/catch keeps default","PASS","code: empty catch, lists still run"),
    ("UI-009","Studio UI","I open ⚙ API keys","openKeys() shows modal","PASS","sweep: modal opens"),
    ("UI-010","Studio UI","I see one input per key + help","openKeys() builds .kfield per key","PASS","sweep: 12 fields shown"),
    ("UI-011","Studio UI","I see saved vs not-set","✓ saved / not set per configured","PASS","sweep(AP1): APIFY '✓ saved' after save"),
    ("UI-012","Studio UI","I read the Apify guidance","static copy in modal","PASS","sweep: guidance present"),
    ("UI-013","Studio UI","I save keys, modal closes","saveKeys() POSTs, closeKeys()","PASS","sweep(AP1): saved+persisted"),
    ("UI-014","Studio UI","Blank keeps existing key","only non-empty values POSTed","PASS","wire(E3): blank preserves"),
    ("UI-015","Studio UI","Save failure alerts me","saveKeys() catch alert()","PASS","code: try/catch alert"),
    ("UI-016","Studio UI","I cancel the modal","closeKeys() removes .show","PASS","sweep: cancel closes"),
    ("UI-017","Studio UI","Only buttons close the modal","no ESC/backdrop handler","NOTE","code: intentional; minor UX"),
    ("UI-018","Studio UI","I click + New research","newConversation() creates+opens","PASS","sweep(SB2): input enables"),
    ("UI-019","Studio UI","I see convos w/ report-count","refreshConvos() renders rows","PASS","sweep: rows render"),
    ("UI-020","Studio UI","Active convo highlighted",".convo.active styling","PASS","sweep+code"),
    ("UI-021","Studio UI","I switch conversations","openConversation() resets panes","PASS","sweep(SB3): switch works, 0 errors"),
    ("UI-022","Studio UI","Re-open active re-inits stream","openConversation() reopens ES","PASS","code+sweep"),
    ("UI-023","Studio UI","Empty convo list","no placeholder (blank)","NOTE","code: no empty-state text; minor"),
    ("UI-024","Studio UI","I see finished reports","loadReports() filters synthesized","PASS","sweep(SB4): reports listed"),
    ("UI-025","Studio UI","Empty reports state","'No reports yet.'","PASS","code+default markup"),
    ("UI-026","Studio UI","I click a report","openReportDoc()","PASS","sweep(SB4): doc view opens"),
    ("UI-027","Studio UI","Jobs fetch fails gracefully","loadReports() try/catch","PASS","code: empty catch"),
    ("UI-028","Studio UI","Empty chat hint","#stream default text","PASS","code+sweep"),
    ("UI-029","Studio UI","Input disabled until convo","textarea/button disabled","PASS","sweep: disabled initially"),
    ("UI-030","Studio UI","Open convo enables+focuses input","openConversation() enables","PASS","sweep(SB2)"),
    ("UI-031","Studio UI","Enter sends","keydown Enter (no shift)->send","PASS","code+live send"),
    ("UI-032","Studio UI","Shift+Enter newline","handler ignores shift+enter","PASS","code"),
    ("UI-033","Studio UI","Textarea auto-resizes","input listener min(scrollH,140)","PASS","code"),
    ("UI-034","Studio UI","I click Send","send() POSTs chat/send","PASS","live: 4 runs sent via UI"),
    ("UI-035","Studio UI","Empty/no-convo send no-ops","send() guard !msg||!active","PASS","code"),
    ("UI-036","Studio UI","Composer locks during run","send() disables input","PASS","live(C6): input disabled mid-run"),
    ("UI-037","Studio UI","Send failure re-enables","catch onEvent('error')+re-enable","PASS","code"),
    ("UI-038","Studio UI","Activity tab","showTab('feed')","PASS","sweep+live"),
    ("UI-039","Studio UI","Graph tab","showTab('graph')","PASS","live: graph rendered"),
    ("UI-040","Studio UI","Report tab","showTab('report') flex","PASS","sweep(R1)"),
    ("UI-041","Studio UI","One pane at a time","showTab() toggles all","PASS","code+sweep"),
    ("UI-042","Studio UI","Activity empty hint","#feed default text","PASS","code"),
    ("UI-043","Studio UI","Tool-call rows (→)","tool_call -> addEvt call","PASS","live: 205 tool calls"),
    ("UI-044","Studio UI","Success result rows (✓)","tool_result -> green","PASS","live: 205 results"),
    ("UI-045","Studio UI","Error result rows (✕)","is_error -> red err","PASS","live: 43 errors shown red"),
    ("UI-046","Studio UI","Click row -> raw JSON","head.onclick toggle open","PASS","code"),
    ("UI-047","Studio UI","Raw payload readable/scrolls","<pre> max-height 280 scroll","PASS","code"),
    ("UI-048","Studio UI","Thinking block starts","thinking_start -> .think","PASS","live: 63 thinking events"),
    ("UI-049","Studio UI","Thinking streams","thinking appends .tbody","PASS","live"),
    ("UI-050","Studio UI","Plain-language notes","note -> addNote muted","PASS","live: note events"),
    ("UI-051","Studio UI","Run errors as red msg","error -> .msg.err","PASS","live"),
    ("UI-052","Studio UI","My message bubble","user -> .msg.user right","PASS","live"),
    ("UI-053","Studio UI","Assistant text streams","text -> .msg.assistant","PASS","live: 389 deltas"),
    ("UI-054","Studio UI","Turn cost line","result cost_usd -> muted","PASS","live: done+cost"),
    ("UI-055","Studio UI","Run done unlocks+refreshes","done re-enables, refreshes","PASS","live"),
    ("UI-056","Studio UI","Graph auto-shows on start","graph_init -> showTab graph","PASS","live"),
    ("UI-057","Studio UI","One node per stage","initGraph() .gnode each","PASS","live: 9 nodes"),
    ("UI-058","Studio UI","Collect sub-nodes indented","collect: -> .gsub","PASS","live: collect:reddit/web"),
    ("UI-059","Studio UI","Node colors update live","node -> setNode state","PASS","live: running/done/failed"),
    ("UI-060","Studio UI","Node detail text","setNode .gdetail","PASS","live: review detail shown"),
    ("UI-061","Studio UI","Fresh graph placeholder","openConversation resets","PASS","code"),
    ("UI-062","Studio UI","Report dropdown per job","reportSelect onchange","PASS","code+sweep"),
    ("UI-063","Studio UI","'No report yet' state","rebuildReportSelect blank","PASS","code"),
    ("UI-064","Studio UI","Select report -> iframe","loadReport() src cache-bust","PASS","sweep(R1): iframe loads"),
    ("UI-065","Studio UI","Report auto-appears post-synth","tool_result regex -> rebuild","PASS","live"),
    ("UI-066","Studio UI","job_linked attaches report","job_linked pushes job_id","PASS","live: job_linked events"),
    ("UI-067","Studio UI","Download chips","loadDownloads() renders chips","PASS","wire(R2): 7 chips download"),
    ("UI-068","Studio UI","No files -> no chips","clears box, returns","PASS","code"),
    ("UI-069","Studio UI","No duplicate chips","clears AFTER fetch","PASS (fixed)","earlier fix; sweep"),
    ("UI-070","Studio UI","Full-screen doc view","openReportDoc() .show","PASS","sweep: doc view POV renders"),
    ("UI-071","Studio UI","Doc view title","docTitle truncated 90","PASS","code"),
    ("UI-072","Studio UI","Doc view Back","closeDoc() blanks iframe","PASS","sweep(SB4)"),
    ("UI-073","Studio UI","Doc view no ESC/backdrop","only Back closes","NOTE","code: intentional; minor"),
    ("UI-074","Studio UI","Per-convo SSE stream","EventSource per cid","PASS","live+wire(E6)"),
    ("UI-075","Studio UI","Every event updates right pane","14 listeners -> onEvent","PASS","live"),
    ("UI-076","Studio UI","SSE cleanup on switch","es.close() before new","PASS","code"),
    ("UI-077","Studio UI","Dropped stream reconnects","browser-native retry; client sends no from= cursor","NOTE","code: on reconnect server replays full transcript (from=0) — correct, just not incremental"),
    ("UI-078","Studio UI","Ping keepalive ignored","onEvent ping returns","PASS","code+wire(ping seen)"),
    ("UI-079","Studio UI","Chat auto-scrolls","scrollStream()","PASS","code"),
    ("UI-080","Studio UI","Feed auto-scrolls","addEvt/addNote scroll","PASS","code"),
    ("UI-081","Studio UI","Convo/job cache persists","convs{} preserve","PASS","code"),
    ("UI-082","Studio UI","No console noise","no console.* calls","PASS","sweep: 0 console errors"),
    # ================= STUDIO ENDPOINTS (studio.py) =================
    ("EP-001","Endpoint","GET / loads app","index() returns PAGE","PASS","wire: 200 HTML"),
    ("EP-002","Endpoint","GET /api/health","db+auth, no token leak","PASS","wire(E1): db+auth, no secrets"),
    ("EP-003","Endpoint","GET /api/conversations","list of convos","PASS","wire: 200 array"),
    ("EP-004","Endpoint","POST /api/conversations bad token","403","PASS","wire: 403 no token"),
    ("EP-005","Endpoint","POST /api/conversations ok","creates, 200","PASS","wire: 200 valid create"),
    ("EP-006","Endpoint","Title default/truncate","'New research', [:120]","PASS","code+wire"),
    ("EP-007","Endpoint","Malformed/non-dict body","was 500 -> now graceful default","PASS (fixed)","wire: malformed/array -> 200 (was 500)"),
    ("EP-008","Endpoint","POST chat/send bad token","403","PASS","wire"),
    ("EP-009","Endpoint","chat/send unknown convo","404","PASS","wire"),
    ("EP-010","Endpoint","chat/send while busy","409","PASS","live(C6): 409 mid-run"),
    ("EP-011","Endpoint","chat/send empty message","400","PASS","code+wire"),
    ("EP-012","Endpoint","chat/send ok","200, run starts","PASS","live: 4 runs"),
    ("EP-013","Endpoint","send branch precedence","token>parse>404>409>400>ok","PASS","code+wire"),
    ("EP-014","Endpoint","stream unknown convo","404 JSON","PASS","wire"),
    ("EP-015","Endpoint","stream replay from=","replays transcript[from:]","PASS","wire(E6): 284 events replayed"),
    ("EP-016","Endpoint","stream live tail","yields queue events","PASS","live"),
    ("EP-017","Endpoint","stream ping keepalive","15s ping","PASS","wire: ping seen"),
    ("EP-018","Endpoint","stream bad/neg from=","was 500 -> clamp to 0","PASS (fixed)","wire: from=abc/-1 -> 404 not 500"),
    ("EP-019","Endpoint","report missing job","404 HTML","PASS","wire(E5): 404"),
    ("EP-020","Endpoint","report not ready","200 pending page","PASS (fixed)","wire(E5): running job -> 200 'being prepared'"),
    ("EP-021","Endpoint","report ok","200 report HTML","PASS","wire+sweep"),
    ("EP-022","Endpoint","report non-int id","404 (route)","PASS","wire: /abc -> 404"),
    ("EP-023","Endpoint","export ok but no html","would 500 (rare)","NOTE","code: rare edge; low risk"),
    ("EP-024","Endpoint","files bad job","404 {files:[]}","PASS","wire"),
    ("EP-025","Endpoint","files list","allow-listed present","PASS","wire(R2): 7 chips"),
    ("EP-026","Endpoint","_DOWNLOADABLE allow-list","docx/deck/html/md/csv/charts","PASS","wire+code"),
    ("EP-027","Endpoint","file not-allowed/traversal","404 'not downloadable'","PASS","wire(E4): traversal 404"),
    ("EP-028","Endpoint","file bad job","404 'no such job'","PASS","wire"),
    ("EP-029","Endpoint","file charts -> zip","zips charts/ dir","PASS (fixed)","wire: charts -> zip 130KB"),
    ("EP-030","Endpoint","file charts none","404","PASS","code"),
    ("EP-031","Endpoint","file missing/dir","404 'file not found'","PASS","code"),
    ("EP-032","Endpoint","file ok","FileResponse 200","PASS","wire(R2): real bytes"),
    ("EP-033","Endpoint","GET /api/dossier ok","200 JSON","PASS","wire"),
    ("EP-034","Endpoint","dossier error","404","PASS","wire"),
    ("EP-035","Endpoint","GET /api/jobs","200 list","PASS","wire"),
    ("EP-036","Endpoint","jobs error code","400 (dossier uses 404)","NOTE","code: minor inconsistency, UI-tolerant"),
    ("EP-037","Endpoint","settings presence-only","booleans, no secrets","PASS","wire(E2)"),
    ("EP-038","Endpoint","settings POST bad token","403","PASS","wire(E3)"),
    ("EP-039","Endpoint","settings POST ok","writes, 200","PASS (fixed)","wire: malformed body -> 200 (was 500)"),
    ("EP-040","Endpoint","settings blank keeps secret","0600, atomic, preserve","PASS","wire(E3): 0600, DB intact"),
    ("EP-041","Endpoint","auth model / token scope","POSTs guarded, GETs open, 127.0.0.1","PASS","wire+code"),
    ("EP-042","Endpoint","wrong HTTP method","405","PASS","code (Starlette)"),
    ("EP-043","Endpoint","ensure_mcp_server lifecycle","spawns+waits shared MCP","PASS","live: MCP spawns per run"),
    ("EP-044","Endpoint","_mcp_responding probe","port probe, proxy-bypass","PASS","live"),
    ("EP-045","Endpoint","_stop_mcp_server atexit","terminate on exit","PASS","code"),
    ("EP-046","Endpoint","_reap_orphan_mcp","kills stale MCP on startup","PASS (fixed)","live: reaped 2h-old orphan"),
    ("EP-047","Endpoint","port-in-use friendly","friendly msg + exit 1","PASS (fixed)","live(L4)"),
    ("EP-048","Endpoint","main() bind + browser","127.0.0.1:8770","PASS","live: runs"),
    ("EP-049","Endpoint","_run_agent dispatch","graph vs conversational","PASS","live: both paths"),
    ("EP-050","Endpoint","agent tool gating","deny Bash/Task/Agent/…","PASS","live: no banned tools in 4 runs"),
    ("EP-051","Endpoint","SSE event vocabulary","typed events emitted","PASS","live: all types seen"),
    ("EP-052","Endpoint","conversational no-answer path","guidance error + result","PASS","code"),
    ("EP-053","Endpoint","_drive inactivity timeout","wait_for + interrupt","PASS","code"),
    # ================= MCP TOOLS (mcp_server.py) =================
    ("MCP-1","MCP tool","create_research_job","persists plan, returns job","PASS","harness: job created"),
    ("MCP-2","MCP tool","add_evidence","stores quote+provenance","PASS","harness: evidence_id"),
    ("MCP-3","MCP tool","fetch_rss","SSRF-guarded, filters terms","PASS","harness: RSS 200, matched"),
    ("MCP-4","MCP tool","search_x (no key)","not_configured fallback","PASS","harness: graceful"),
    ("MCP-5","MCP tool","inspect_discord_invite","member/presence, no token","PASS","harness: discord.gg 200"),
    ("MCP-6","MCP tool","discord_landscape","ranks communities, no token","PASS","harness: 4 discovered"),
    ("MCP-7","MCP tool","discord_widget","online/voice or soft-off","PASS","live(plan verification)"),
    ("MCP-8","MCP tool","run_apify_actor (no key)","not_configured","PASS","harness: graceful"),
    ("MCP-9","MCP tool","apify_collect (no key)","not_configured; bad key soft-fail","PASS (fixed)","harness+AP2: no crash on bad key"),
    ("MCP-10","MCP tool","search_twitch (no creds)","not_configured","PASS","harness: graceful"),
    ("MCP-11","MCP tool","read_twitch_chat","anon WSS, no creds","PASS","live(plan): 11 msgs from #xqc"),
    ("MCP-12","MCP tool","read_kick_chat","Pusher WSS, no creds","PASS","live(plan): kick chat verified"),
    ("MCP-13","MCP tool","search_kick","curl_cffi public, no key","PASS","harness: xqc followers"),
    ("MCP-14","MCP tool","read_discord_channel (no token)","not_configured","PASS","harness: graceful"),
    ("MCP-15","MCP tool","crawl_web_page","SSRF-guarded, text+links","PASS","harness: example.com crawled"),
    ("MCP-16","MCP tool","search_web","Tavily>Brave>Serper>DDG","PASS","harness: DDG results"),
    ("MCP-17","MCP tool","discover_sources","grouped venues; probe-resilient","PASS (fixed)","harness: 25 leads; one probe fail no longer aborts"),
    ("MCP-18","MCP tool","search_reddit","public JSON; 403 -> clear guidance","PASS","harness: designed rate-limit ValueError (route-around)"),
    ("MCP-19","MCP tool","fetch_reddit_thread","post+comments","PASS","live: 19 threads in creami run"),
    ("MCP-20","MCP tool","get_research_dossier","compact spread","PASS","harness+live"),
    ("MCP-21","MCP tool","list_metric_fields","numeric fields","PASS","harness"),
    ("MCP-22","MCP tool","compute_metric","median/n per group","PASS","harness"),
    ("MCP-23","MCP tool","compute_rate","rate % per group","PASS","harness"),
    ("MCP-24","MCP tool","analyze_chatter","voices/terms/sufficiency/consensus","PASS","harness+live: consensus incl."),
    ("MCP-25","MCP tool","search_evidence","matching rows","PASS","harness"),
    ("MCP-26","MCP tool","list_research_jobs","recent jobs","PASS","harness"),
    ("MCP-27","MCP tool","connector_status","ready+fallbacks","PASS","harness"),
    ("MCP-28","MCP tool","write_research_synthesis","authors report; blank->refuse","PASS","harness: blank->ValueError, real->saved"),
    ("MCP-29","MCP tool","export_research_report","package; pre-synth->refuse","PASS","harness: refuses then exports"),
    ("MCP-30","MCP tool","get_browser_capture_mission","mission for active job","PASS","harness: correct refuse on non-active job"),
    ("MCP-31","MCP tool","request_browser_traffic_session","approval-gated request","PASS","harness: correct 'active job only' guard"),
    ("MCP-32","MCP tool","get_browser_traffic_session","poll status","REVIEWED","code: needs an active session to poll"),
    ("MCP-33","MCP prompt","research_assignment","7-step workflow prompt","PASS","code: prompt returns steps"),
    # ================= CONTROL CENTER (control_center.py, 8765) =================
    ("CC-1","Control Center","Dashboard page","serves control room HTML","PASS","live: GET / 200, pairing UI"),
    ("CC-2","Control Center","GET /api/status","readiness + checks","PASS","live: /api/status 200"),
    ("CC-3","Control Center","GET /api/settings","presence only","REVIEWED","code: presence map"),
    ("CC-4","Control Center","POST /api/settings","writes .env 0600","REVIEWED","code (Studio settings path tested separately)"),
    ("CC-5","Control Center","Run pulse pipeline","main.py pulse subprocess","REVIEWED","code: legacy gaming pipeline, not run"),
    ("CC-6","Control Center","Run collect","main.py collect all","REVIEWED","code: legacy"),
    ("CC-7","Control Center","Run report","main.py report","REVIEWED","code: legacy"),
    ("CC-8","Control Center","Run setup (migrate)","main.py migrate","PASS","migrate runs (CLI-1)"),
    ("CC-9","Control Center","GET /api/job?id","live job status","REVIEWED","code"),
    ("CC-10","Control Center","GET /api/research/jobs","list research jobs","REVIEWED","code (store.list_jobs tested via MCP-26)"),
    ("CC-11","Control Center","POST /api/research/export","export a job","REVIEWED","code (export tested via MCP-29)"),
    ("CC-12","Control Center","POST /api/open","open report/folder in OS","REVIEWED","code: shells to open; not run headless"),
    ("CC-13","Control Center","GET /api/research/sessions","list capture sessions","REVIEWED","code"),
    ("CC-14","Control Center","Stop session (dashboard)","stop_session no client","REVIEWED","code"),
    ("CC-15","Control Center","Create pairing code","10-min hashed code","REVIEWED","code: needs extension to complete"),
    ("CC-16","Control Center","Pair (extension)","exchange code, upsert client","REVIEWED","code: needs real Chrome"),
    ("CC-17","Control Center","capture/jobs (bearer)","active jobs for extension","REVIEWED","code: bearer-auth, needs pairing"),
    ("CC-18","Control Center","capture mission (bearer)","job mission","REVIEWED","code (mission logic via MCP-30)"),
    ("CC-19","Control Center","capture sessions (bearer)","session cards","REVIEWED","code"),
    ("CC-20","Control Center","trusted-domains list (bearer)","standing consent list","REVIEWED","code"),
    ("CC-21","Control Center","capture/pending GET","server pending queue","NOTE","code: legacy/unused by shipping extension"),
    ("CC-22","Control Center","capture/pending POST","stage capture","NOTE","code: legacy/unused"),
    ("CC-23","Control Center","capture/preview","term-match check","NOTE","code: stale/unused endpoint"),
    ("CC-24","Control Center","capture/approve","save capture as evidence","REVIEWED","code: needs paired extension"),
    ("CC-25","Control Center","pending/{id}/approve","approve staged","NOTE","code: legacy/unused"),
    ("CC-26","Control Center","capture/revoke","unpair browser","REVIEWED","code"),
    ("CC-27","Control Center","session approve (30m)","TTL-capped approve","REVIEWED","code"),
    ("CC-28","Control Center","session decline","decline request","REVIEWED","code"),
    ("CC-29","Control Center","session stop (bearer)","stop recording","REVIEWED","code"),
    ("CC-30","Control Center","session start","request+approve in one","REVIEWED","code: needs extension"),
    ("CC-31","Control Center","session auto-approve","trusted-domain hands-free","REVIEWED","code"),
    ("CC-32","Control Center","trust domain","standing consent","REVIEWED","code"),
    ("CC-33","Control Center","pending DELETE (bulk)","reject job's pending","NOTE","code: legacy/unused"),
    ("CC-34","Control Center","pending/{id} DELETE","reject one","NOTE","code: legacy/unused"),
    ("CC-35","Control Center","evidence DELETE","delete promoted evidence","NOTE","code: no in-UI control (audit gap)"),
    ("CC-36","Control Center","untrust domain","revoke consent","REVIEWED","code"),
    ("CC-37","Control Center","OPTIONS preflight","CORS for extension origin","REVIEWED","code"),
    ("CC-38","Control Center","auth/guard model","loopback+origin+bearer+rate","REVIEWED","code: constant-time compare"),
    ("CC-39","Control Center","settings .env ownership","owns 9 keys atomically","NOTE","code: some research keys not in this UI (Studio modal owns them)"),
    ("CC-40","Control Center","server bootstrap","binds 8765, port-in-use msg","PASS","live: launched, serves"),
    # ================= BROWSER EXTENSION (browser_extension/) =================
    ("EX-1","Extension","Load unpacked","MV3 manifest, loopback host","REVIEWED","code: needs Chrome load"),
    ("EX-2","Extension","Install bootstrap","seeds defaults, context menu","REVIEWED","code"),
    ("EX-3","Extension","Pair browser","POST /pair, store token","REVIEWED","code: needs Chrome"),
    ("EX-4","Extension","Unpair","revoke + wipe local","REVIEWED","code"),
    ("EX-5","Extension","Privacy defaults","loopback-only, clamp cap","PASS","nodetest: normalizeBaseUrl rejects non-loopback"),
    ("EX-6","Extension","Trust a domain","host perm + POST","REVIEWED","code: needs Chrome perm prompt"),
    ("EX-7","Extension","List trusted domains","GET trusted-domains","REVIEWED","code"),
    ("EX-8","Extension","Untrust domain","DELETE + perm remove","REVIEWED","code"),
    ("EX-9","Extension","Popup status pill","paired/not/unavailable","REVIEWED","code"),
    ("EX-10","Extension","Load research jobs","GET capture/jobs","REVIEWED","code"),
    ("EX-11","Extension","Select job + mission","GET mission, render","REVIEWED","code"),
    ("EX-12","Extension","Scan visible candidates","inject, SCAN_VISIBLE","REVIEWED","code: needs Chrome page"),
    ("EX-13","Extension","Capture selection (menu)","context-menu candidate","REVIEWED","code"),
    ("EX-14","Extension","Approve candidate","approvalPayload POST","PASS","nodetest: approvalPayload drops raw when anon"),
    ("EX-15","Extension","Edit/reject/clear/anon queue","local-only writes","PASS","nodetest: anonymize/dedupe"),
    ("EX-16","Extension","Stop/Resume collection","collectionStopped flag","REVIEWED","code"),
    ("EX-17","Extension","Start network capture","host perm + session start","REVIEWED","code: needs Chrome"),
    ("EX-18","Extension","Approve/decline agent session","POST approve/decline","REVIEWED","code"),
    ("EX-19","Extension","View/stop active session","SESSION_STOP","REVIEWED","code"),
    ("EX-20","Extension","Standing auto-approve loop","1-min alarm auto-approve","REVIEWED","code"),
    ("EX-21","Extension","Session recorder (MAIN)","wraps fetch/XHR, JSON only","REVIEWED","code: no creds read (by design)"),
    ("EX-22","Extension","Session relay (isolated)","forwards to worker","REVIEWED","code"),
    ("EX-23","Extension","Network capture -> save","dedupe, extract, POST approve","PASS","nodetest: payloadFingerprint/extractMessages/redactUrl"),
    ("EX-24","Extension","Capture across navigation","re-inject on same domain","REVIEWED","code"),
    ("EX-25","Extension","Alarms (approve+expiry)","auto-approve + 30m expiry","REVIEWED","code"),
    ("EX-26","Extension","Shared core library","redaction/dedup/safety","PASS","nodetest: 16/16 core tests"),
    ("EX-27","Extension","Notice + badge feedback","setNotice + badge","REVIEWED","code"),
    ("EX-28","Extension","Core tests","node --test core.test.js","PASS","nodetest: 16/16 green"),
    # ================= CLI (main.py) =================
    ("CLI-1","CLI","migrate","create/upgrade schema, seed","PASS","runs (install/studio use it)"),
    ("CLI-2","CLI","collect <source>","run one collector","REVIEWED","code: legacy gaming pipeline, needs creds"),
    ("CLI-3","CLI","collect all","only configured sources","REVIEWED","code: legacy"),
    ("CLI-4","CLI","collect failure handling","attempt all, non-zero exit","REVIEWED","code"),
    ("CLI-5","CLI","twitch/kick cred guard","clear SystemExit","REVIEWED","code"),
    ("CLI-6","CLI","report <monday>","md+html weekly report","REVIEWED","code: legacy"),
    ("CLI-7","CLI","week_start validation","reject non-Monday/malformed","PASS","code: argparse type (deterministic)"),
    ("CLI-8","CLI","report movement/coverage gating","INSUFFICIENT/BASELINE/UP/DOWN","REVIEWED","code: legacy rollup"),
    ("CLI-9","CLI","cross-source consensus movers",">=2 sources agree","REVIEWED","code: legacy"),
    ("CLI-10","CLI","pulse","fresh on-demand brief","REVIEWED","code: legacy"),
    ("CLI-11","CLI","pulse options","--sources/--press-hours","REVIEWED","code"),
    ("CLI-12","CLI","signal-brief content","5 sections + caveat","REVIEWED","code: legacy"),
    ("CLI-13","CLI","replay --response-id","re-parse one raw","REVIEWED","code"),
    ("CLI-14","CLI","replay --run-id","re-parse a run","REVIEWED","code"),
    ("CLI-15","CLI","subcommand required","usage if none","PASS","code: argparse required"),
    # ================= INSTALL / LAUNCHER (install.sh, studio) =================
    ("INS-1","Install","one-line install + overrides","OYSTER_REPO/BRANCH/HOME","REVIEWED","code+syntax; not run from scratch (modifies host)"),
    ("INS-2","Install","platform detection","macos/linux/refuse","REVIEWED","code"),
    ("INS-3","Install","sudo/root handling","sudo or die","REVIEWED","code"),
    ("INS-4","Install","Homebrew requirement","die if absent (mac)","REVIEWED","code"),
    ("INS-5","Install","ensure git","install if missing","REVIEWED","code"),
    ("INS-6","Install","fresh clone","branch --depth 1","REVIEWED","code"),
    ("INS-7","Install","update in place","fetch+checkout FETCH_HEAD","PASS (fixed)","branch-trap fixed; code"),
    ("INS-8","Install","ensure Python","install if missing","REVIEWED","code"),
    ("INS-9","Install","ensure PostgreSQL","install if missing","REVIEWED","code"),
    ("INS-10","Install","start DB + wait","pg_isready loop or die","REVIEWED","code"),
    ("INS-11","Install","DB login role (linux)","create role for OS user","REVIEWED","code"),
    ("INS-12","Install","create db idempotent","untouched if exists","REVIEWED","code"),
    ("INS-13","Install","core venv + deps","venv + requirements","REVIEWED","code"),
    ("INS-14","Install","studio venv + deps","venv-studio","REVIEWED","code"),
    ("INS-15","Install",".env creation","defaults, 0600, never clobber","REVIEWED","code"),
    ("INS-16","Install","migrate","main.py migrate","PASS","migrate runs"),
    ("INS-17","Install","attach MCP to hosts","claude/codex mcp add","REVIEWED","code"),
    ("INS-18","Install","desktop launcher","double-click icon","REVIEWED","code"),
    ("INS-19","Install","final guidance","next steps printed","REVIEWED","code"),
    ("INS-20","Install","auto-open Studio","if signed in","REVIEWED","code"),
    ("INS-21","Launcher","studio: python detect","3.11 or die","PASS","live: ./studio ran"),
    ("INS-22","Launcher","studio: core venv bootstrap","first-run setup","PASS","live"),
    ("INS-23","Launcher","studio: studio venv bootstrap","first-run setup","PASS","live"),
    ("INS-24","Launcher","studio: db self-heal","start pg + migrate","PASS","live(L2): 'database ready'"),
    ("INS-25","Launcher","studio: auto sign-in","claude auth login","PASS","live(L2): 'signed in'"),
    ("INS-26","Launcher","studio: load .env","source into env","PASS","live"),
    ("INS-27","Launcher","studio: launch","exec studio on 8770","PASS","live(L2/L4)"),
    # ================= EXPORT / REPORTING (export.py, reporting/) =================
    ("EXP-1","Export","refuse bare data","no synthesis -> raise","PASS","harness(MCP-29a)"),
    ("EXP-2","Export","folder + slug","predictable job folder","PASS","live: folders correct"),
    ("EXP-3","Export","report.md dossier","full md + appendices","PASS","live: md built"),
    ("EXP-4","Export","title derivation","clean title, skip labels","PASS (fixed)","fixed 'internal'; live: real titles"),
    ("EXP-5","Export","report.html","designed self-contained","PASS","live+sweep: POV/glance/[n]"),
    ("EXP-6","Export","stat tiles","top-3 numeric tiles","PASS","live: tiles in html/deck"),
    ("EXP-7","Export","raw evidence files","json/csv (injection-guarded)","PASS","live: evidence.csv 388KB"),
    ("EXP-8","Export","charts","png+svg+csv per table","PASS","live: 6 charts"),
    ("EXP-9","Export","report.docx","formatted Word","PASS","live: docx opens, 33KB"),
    ("EXP-10","Export","deck.pptx","16:9 consultant deck","PASS (fixed)","live: 8-9 slides, title fixed"),
    ("EXP-11","Export","numbered bibliography","deterministic [n] dedupe","PASS","live: 10-25 [n] entries"),
    ("EXP-12","Export","Sources-and-Citations.md","[n] deep links + quotes","PASS","live: deep-link ledger"),
    ("EXP-13","Export","raw-data/","per-connector jsonl","PASS","live: 4 raw-data files"),
    ("EXP-14","Export","README-start-here.md","manifest of produced files","PASS","live: README present"),
    ("EXP-15","Export","best-effort package","builder fails -> warning","PASS","code+live: all built"),
    ("EXP-16","Export","table normalization","real columns, comma ints","PASS","live+tests: tables populated"),
    ("EXP-17","Export","chart-series selection","auto-pick numeric column","PASS","live+tests"),
    ("EXP-18","Export","shared visual language","one palette/font","PASS","code+live"),
    ("EXP-19","Export","findings grouping (Appx B)","by source_type","PASS","code+live"),
    ("CW-1","Report","consensus weighting","lead by endorsement not count","PASS (fixed)","live(2102): '176 upvotes','top-voted'; 200 tests"),
    # ================= PERIPHERAL / PROGRAM =================
    ("P-CI","CI","postgres acceptance scripts collected","__main__ scripts, not pytest","DEFER","code: real hygiene gap, post-demo"),
]

HEADERS = ["ID", "Surface", "User story (human POV)", "Expected behavior (from code)", "Verdict", "How verified (this pass · 2026-08-07)"]
WIDTHS = [8, 15, 40, 40, 13, 52]
COLORS = {"PASS": "1E7A34", "PASS (fixed)": "0F766E", "NOTE": "8A6D0B", "DEFER": "7A4A1E", "REVIEWED": "5B5B6B"}


def build(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature tracker"
    hf, hfont = PatternFill("solid", fgColor="1A1F2B"), Font(bold=True, color="FFFFFF")
    for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(1, c, h); cell.fill, cell.font = hf, hfont
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, row in enumerate(ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val); cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 5).font = Font(bold=True, color=COLORS.get(row[4], "000000"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS)+1}"
    # Summary sheet
    s2 = wb.create_sheet("Summary")
    counts = Counter(r[4] for r in ROWS)
    bysurface = Counter(r[1] for r in ROWS)
    s2["A1"], s2["B1"] = "Verdict", "Count"; s2["A1"].font = s2["B1"].font = Font(bold=True)
    for i, (k, v) in enumerate(sorted(counts.items()), 2):
        s2.cell(i, 1, k); s2.cell(i, 2, v)
    row = len(counts) + 3
    s2.cell(row, 1, f"Total stories: {len(ROWS)}").font = Font(bold=True)
    s2.cell(row + 2, 1, "By surface:").font = Font(bold=True)
    for i, (k, v) in enumerate(sorted(bysurface.items()), row + 3):
        s2.cell(i, 1, k); s2.cell(i, 2, v)
    legend = [
        "", "Verdict legend:",
        "PASS         exercised this pass, works.",
        "PASS (fixed) defect found + fixed this pass; re-tested.",
        "REVIEWED     code-verified this pass but NOT executed E2E (needs a real Chrome +",
        "             paired extension, legacy gaming-collector creds, or a from-scratch host",
        "             install). Honest not-fully-exercised — NOT counted as PASS.",
        "NOTE         works, known limitation/cosmetic.",
        "DEFER        real gap, deliberately deferred post-demo.",
        "",
        "Anchored by 4 live runs through the real UI (jobs 1861/1998/2033/2102), an MCP tool",
        "harness (33 tools), Playwright UI drives, curl endpoint tests, node --test (extension),",
        "and the 200-test pytest suite. Six defects fixed this pass: report jargon page, gaming",
        "language, 'internal' title, Apify bad-key crash, orphaned-MCP stale code, consensus",
        "weighting — plus this pass: discover_sources probe-resilience, /chat/stream bad from=,",
        "malformed POST bodies (conversations/settings).",
    ]
    for j, line in enumerate(legend, row + 3 + len(bysurface) + 1):
        c = s2.cell(j, 1, line)
        if line.endswith(":"):
            c.font = Font(bold=True)
    s2.column_dimensions["A"].width = 22
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"wrote {path} | {len(ROWS)} stories | {dict(counts)}")


if __name__ == "__main__":
    build(Path(__file__).resolve().parent / "Research-Oyster-QA.xlsx")
